#!/usr/bin/env python3
"""Build Northline_Capex_Investment_Case.xlsx — plant capex NPV / IRR / payback pack."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.axis import NumericAxis
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

yellow = PatternFill("solid", fgColor="FFF2CC")
header_fill = PatternFill("solid", fgColor="1F4E79")
section_fill = PatternFill("solid", fgColor="D6E3F0")
green_fill = PatternFill("solid", fgColor="C6EFCE")
amber_fill = PatternFill("solid", fgColor="FFE699")
red_fill = PatternFill("solid", fgColor="F8CBAD")
tile_fill = PatternFill("solid", fgColor="E9EDF4")
light_gray = PatternFill("solid", fgColor="F5F5F5")
input_font = Font(name="Calibri", size=11, color="0000FF")
black = Font(name="Calibri", size=11, color="000000")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
title_font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
section_font = Font(name="Calibri", size=12, bold=True, color="1F4E79")
bold = Font(name="Calibri", size=11, bold=True)
bold_black = Font(name="Calibri", size=11, bold=True, color="000000")
italic_grey = Font(name="Calibri", size=10, italic=True, color="666666")
small_grey = Font(name="Calibri", size=9, italic=True, color="666666")
link_font = Font(name="Calibri", size=11, color="0563C1", underline="single")
tile_label = Font(name="Calibri", size=9, color="666666")
tile_value = Font(name="Calibri", size=14, bold=True, color="1F4E79")
thin = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
money = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
pct = "0.0%"
num = "#,##0.0"
ppu = "0.00"
int_fmt = "#,##0"
years_fmt = "0.0"

YEARS = list(range(2026, 2035))  # 2026-2034, t = 0..8
N = len(YEARS)

ASSUMP = "01_Assumptions"
CF = "02_Cash_Flows"
RET = "03_Returns"
SENS = "04_Sensitivity"
ONE = "05_One_Pager"


def style_input(cell):
    cell.fill = yellow
    cell.font = input_font
    cell.border = thin
    cell.alignment = Alignment(horizontal="center")


def style_formula(cell, key=False):
    cell.font = bold_black if key else black
    cell.border = thin
    cell.alignment = Alignment(horizontal="center")
    if key:
        cell.fill = green_fill


def style_header_cell(cell, value):
    cell.value = value
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def shade_section(ws, row, until=12):
    ws.cell(row, 2).fill = section_fill
    ws.cell(row, 2).font = section_font
    for c in range(3, until + 1):
        ws.cell(row, c).fill = section_fill


def landscape(ws, fit_height=0):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = fit_height
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
    ws.sheet_view.showGridLines = False
    ws.page_setup.horizontalCentered = True
    ws.oddFooter.left.text = "Northline Consumer Products  |  fictional sample  |  $000s"
    ws.oddFooter.right.text = "Page &P of &N"


def ycol(i):
    return get_column_letter(3 + i)


wb = Workbook()
wb.properties.creator = "Sai Siri Bandaru"
wb.properties.title = "Northline Capex Investment Case"

# ========== 00_Cover ==========
ws = wb.active
ws.title = "00_Cover"
ws.sheet_properties.tabColor = "1F4E79"
set_col_widths(ws, [4, 96])
landscape(ws, fit_height=1)

ws["B2"] = "Northline Consumer Products"
ws["B2"].font = title_font
ws["B3"] = "Capex investment case  ·  Plant 2 RTD Line 4  ·  NPV, IRR, payback"
ws["B3"].font = section_font
ws["B4"] = "FY2026–2034  ·  unlevered project cash flows  ·  fictional CPG company  ·  $000s"
ws["B4"].font = italic_grey

ws["B6"] = "What this file is"
ws["B6"].font = bold
ws["B7"] = (
    "An investment-committee pack for a fictional high-speed RTD filling line at Plant 2. "
    "Capex is timed over two years. Benefits are incremental volume × contribution per case "
    "plus run-rate cost savings (labor, scrap, energy, avoided maintenance on the old line). "
    "Cash flows are unlevered: NOPAT + D&A − capex − ΔNWC + after-tax salvage. "
    "Returns (NPV, IRR, payback, PI) and a WACC × volume sensitivity sit on formulas. "
    "A Base / Upside / Downside toggle scales capex, volume, savings, and contribution."
)
ws["B7"].alignment = Alignment(wrap_text=True)
ws.row_dimensions[7].height = 72

ws["B9"] = "How to use"
ws["B9"].font = bold
ws["B10"] = "1. Edit yellow cells on 01_Assumptions (scenario, WACC, tax, life, capex timing, volume ramp, unit economics, savings)."
ws["B11"] = "2. Read the unlevered cash-flow build on 02_Cash_Flows (all formulas from Assumptions)."
ws["B12"] = "3. Read NPV, IRR, payback, and PI on 03_Returns."
ws["B13"] = "4. Stress WACC and volume on 04_Sensitivity (one-way, two-way, tornado)."
ws["B14"] = "5. Use 05_One_Pager as the investment-committee snapshot."

ws["B16"] = "File conventions"
ws["B16"].font = bold
ws["B17"] = "Yellow cells with blue font = inputs. Black font = formulas. Green cells = key outputs."
ws["B17"].fill = yellow
ws["B17"].font = Font(name="Calibri", size=11, color="0000FF", bold=True)
ws["B18"] = "Discounting is year-end. D&A is straight-line on (gross capex − salvage) over project life, starting Year 1 (COD)."
ws["B19"] = "All figures are fictional. There is no employer data in this file. Dollar figures in $000s. Volume in 000 cases."

ws["B21"] = "Portfolio"
ws["B21"].font = bold
ws["B22"] = "Sai Siri Bandaru — Financial Analyst | FP&A | forecasting, variance analysis, Excel"
ws["B23"] = "https://github.com/saisiri-bandaru"
ws["B23"].font = link_font

# ========== 01_Assumptions ==========
ws = wb.create_sheet(ASSUMP)
ws.sheet_properties.tabColor = "F7C948"
set_col_widths(ws, [4, 42, 14, 14, 14, 14, 14, 14, 14, 14, 14, 18])
landscape(ws)
ws.freeze_panes = "C6"

ws["B2"] = "Assumptions"
ws["B2"].font = title_font
ws["B3"] = "Yellow + blue = inputs. Scenario, WACC, capex timing, volume ramp, and unit economics drive every other tab."
ws["B3"].font = italic_grey

ws["B5"] = "Control panel"
shade_section(ws, 5, until=6)
ws["B6"] = "Scenario (1 = Base, 2 = Upside, 3 = Downside)"
ws["C6"] = 1
style_input(ws["C6"])
ws["C6"].number_format = "0"
ws["D6"] = '=INDEX({"Base","Upside","Downside"},1,C6)'
style_formula(ws["D6"], key=True)
ws["E6"] = "Scales capex, volume, savings, and contribution $/case."
ws["E6"].font = small_grey

ws["B7"] = "Capex factor"
ws["C7"] = "=INDEX($D$15:$D$17,$C$6)"
style_formula(ws["C7"])
ws["C7"].number_format = "0.000"
ws["B8"] = "Volume factor"
ws["C8"] = "=INDEX($E$15:$E$17,$C$6)"
style_formula(ws["C8"])
ws["C8"].number_format = "0.000"
ws["B9"] = "Savings factor"
ws["C9"] = "=INDEX($F$15:$F$17,$C$6)"
style_formula(ws["C9"])
ws["C9"].number_format = "0.000"
ws["B10"] = "Contribution factor (on $/case)"
ws["C10"] = "=INDEX($G$15:$G$17,$C$6)"
style_formula(ws["C10"])
ws["C10"].number_format = "0.000"

ws["B12"] = "Scenario table (inputs)"
shade_section(ws, 12, until=8)
for i, h in enumerate(["#", "Name", "Capex", "Volume", "Savings", "Contrib $/case"]):
    style_header_cell(ws.cell(13, 2 + i), h)
scenarios = [
    (1, "Base", 1.00, 1.00, 1.00, 1.00),
    (2, "Upside", 0.95, 1.12, 1.08, 1.03),
    (3, "Downside", 1.12, 0.85, 0.90, 0.97),
]
for i, rowv in enumerate(scenarios):
    r = 15 + i
    ws.cell(r, 2, rowv[0]).border = thin
    ws.cell(r, 3, rowv[1]).border = thin
    for j, v in enumerate(rowv[2:]):
        cell = ws.cell(r, 4 + j, v)
        style_input(cell)
        cell.number_format = "0.000"

dv = DataValidation(type="list", formula1="1,2,3", allow_blank=False)
dv.error = "Enter 1, 2, or 3"
dv.errorTitle = "Scenario"
dv.prompt = "1 Base / 2 Upside / 3 Downside"
dv.promptTitle = "Scenario"
ws.add_data_validation(dv)
dv.add(ws["C6"])

ws["B19"] = "Project"
shade_section(ws, 19, until=4)
ws["B20"] = "Project name"
ws["C20"] = "Plant 2 — RTD Line 4 (high-speed filling)"
style_input(ws["C20"])
ws["C20"].alignment = Alignment(horizontal="left")
ws.merge_cells("C20:G20")
ws["B21"] = "Sponsor / site"
ws["C21"] = "Operations / Supply Chain  ·  Midwest Plant 2"
style_input(ws["C21"])
ws["C21"].alignment = Alignment(horizontal="left")
ws.merge_cells("C21:G21")
ws["B22"] = "In-service (COD) year"
ws["C22"] = 2027
style_input(ws["C22"])
ws["C22"].number_format = "0"

ws["B24"] = "Returns, tax, NWC, salvage"
shade_section(ws, 24, until=5)
ws["B25"] = "WACC"
ws["C25"] = 0.095
style_input(ws["C25"])
ws["C25"].number_format = pct
ws["D25"] = "Unlevered project hurdle. Sample 9.5%."
ws["D25"].font = small_grey
ws["B26"] = "Tax rate"
ws["C26"] = 0.25
style_input(ws["C26"])
ws["C26"].number_format = pct
ws["B27"] = "Project life (years)"
ws["C27"] = 8
style_input(ws["C27"])
ws["C27"].number_format = "0"
ws["D27"] = "D&A and operating life from COD (t = 1..life). Sample through 2034."
ws["D27"].font = small_grey
ws["B28"] = "Salvage % of gross capex"
ws["C28"] = 0.08
style_input(ws["C28"])
ws["C28"].number_format = pct
ws["B29"] = "NWC % of incremental revenue"
ws["C29"] = 0.065
style_input(ws["C29"])
ws["C29"].number_format = pct
ws["B30"] = "Incremental G&A ($000s / year from COD)"
ws["C30"] = 80
style_input(ws["C30"])
ws["C30"].number_format = money

ws["B32"] = "Year-by-year drivers"
shade_section(ws, 32, until=11)
style_header_cell(ws.cell(33, 2), "Driver")
for i, y in enumerate(YEARS):
    style_header_cell(ws.cell(33, 3 + i), str(y))

ws["B34"] = "Time index t"
for i in range(N):
    cell = ws.cell(34, 3 + i, i)
    style_formula(cell)
    cell.number_format = "0"

# Volume ramp % of run-rate
ws["B35"] = "Volume ramp (% of run-rate)"
vol_ramp = [0.00, 0.42, 0.85, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00]
for i, v in enumerate(vol_ramp):
    cell = ws.cell(35, 3 + i, v)
    style_input(cell)
    cell.number_format = pct

ws["B36"] = "Savings ramp (% of run-rate)"
sav_ramp = [0.00, 0.50, 0.90, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00]
for i, v in enumerate(sav_ramp):
    cell = ws.cell(36, 3 + i, v)
    style_input(cell)
    cell.number_format = pct

# Capex spend pre-factor $000s
ws["B37"] = "Equipment & long-lead ($000s)"
equip = [11800, 2200, 0, 0, 0, 0, 0, 0, 0]
for i, v in enumerate(equip):
    cell = ws.cell(37, 3 + i, v)
    style_input(cell)
    cell.number_format = money

ws["B38"] = "Installation / commissioning ($000s)"
install = [2800, 2800, 0, 0, 0, 0, 0, 0, 0]
for i, v in enumerate(install):
    cell = ws.cell(38, 3 + i, v)
    style_input(cell)
    cell.number_format = money

ws["B39"] = "Spares / startup ($000s)"
spares = [1600, 800, 0, 0, 0, 0, 0, 0, 0]
for i, v in enumerate(spares):
    cell = ws.cell(39, 3 + i, v)
    style_input(cell)
    cell.number_format = money

ws["B40"] = "Gross capex (pre-factor)"
for i in range(N):
    c = ycol(i)
    cell = ws.cell(40, 3 + i, f"={c}37+{c}38+{c}39")
    style_formula(cell)
    cell.number_format = money

ws["B41"] = "Gross capex × scenario"
for i in range(N):
    c = ycol(i)
    cell = ws.cell(41, 3 + i, f"={c}40*$C$7")
    style_formula(cell, key=True)
    cell.number_format = money

ws["B42"] = "Maintenance capex ($000s)"
maint = [0, 0, 180, 180, 190, 190, 200, 200, 210]
for i, v in enumerate(maint):
    cell = ws.cell(42, 3 + i, v)
    style_input(cell)
    cell.number_format = money

ws["B43"] = "Volume (000 cases)"
for i in range(N):
    c = ycol(i)
    cell = ws.cell(43, 3 + i, f"=$C$47*{c}35*$C$8")
    style_formula(cell, key=True)
    cell.number_format = int_fmt

ws["B44"] = "Incremental G&A ($000s)"
for i in range(N):
    c = ycol(i)
    # G&A from COD (t>=1)
    cell = ws.cell(44, 3 + i, f'=IF({c}34>=1,$C$30,0)')
    style_formula(cell)
    cell.number_format = money

for r in range(34, 45):
    ws.cell(r, 2).border = thin
    if r in (40, 41, 43):
        ws.cell(r, 2).font = bold

ws["B46"] = "Unit economics and run-rate savings"
shade_section(ws, 46, until=5)
ws["B47"] = "Run-rate volume lift (000 cases / year)"
ws["C47"] = 1500
style_input(ws["C47"])
ws["C47"].number_format = int_fmt
ws["D47"] = "Incremental cases vs doing nothing. Sample ~Line 4 nameplate after ramp."
ws["D47"].font = small_grey
ws["B48"] = "Incremental net ASP ($ / case)"
ws["C48"] = 11.75
style_input(ws["C48"])
ws["C48"].number_format = ppu
ws["B49"] = "Incremental contribution ($ / case)"
ws["C49"] = 3.20
style_input(ws["C49"])
ws["C49"].number_format = ppu
ws["D49"] = "Net price − variable COGS on the incremental volume. Then × contribution factor."
ws["D49"].font = small_grey
ws["B50"] = "Incremental distribution ($ / case)"
ws["C50"] = 0.42
style_input(ws["C50"])
ws["C50"].number_format = ppu

ws["B52"] = "Direct labor savings — run-rate ($000s)"
ws["C52"] = 2500
style_input(ws["C52"])
ws["C52"].number_format = money
ws["B53"] = "Scrap / giveaway savings — run-rate ($000s)"
ws["C53"] = 800
style_input(ws["C53"])
ws["C53"].number_format = money
ws["B54"] = "Energy savings — run-rate ($000s)"
ws["C54"] = 400
style_input(ws["C54"])
ws["C54"].number_format = money
ws["B55"] = "Avoided old-line maintenance — run-rate ($000s)"
ws["C55"] = 250
style_input(ws["C55"])
ws["C55"].number_format = money
ws["B56"] = "Total run-rate savings ($000s)"
ws["C56"] = "=SUM(C52:C55)"
style_formula(ws["C56"], key=True)
ws["C56"].number_format = money
ws["D56"] = "Applied × savings ramp × savings factor on the cash-flow sheet."
ws["D56"].font = small_grey

ws["B58"] = "Depreciation (straight-line)"
shade_section(ws, 58, until=5)
ws["B59"] = "Total gross capex (post-factor)"
ws["C59"] = "=SUM(C41:K41)"
style_formula(ws["C59"], key=True)
ws["C59"].number_format = money
ws["B60"] = "Salvage value (end of life)"
ws["C60"] = "=C59*C28"
style_formula(ws["C60"])
ws["C60"].number_format = money
ws["B61"] = "Depreciable basis"
ws["C61"] = "=C59-C60"
style_formula(ws["C61"])
ws["C61"].number_format = money
ws["B62"] = "Annual D&A (t = 1..life)"
ws["C62"] = "=IF(C27=0,0,C61/C27)"
style_formula(ws["C62"], key=True)
ws["C62"].number_format = money
ws["D62"] = "Book value at end = salvage, so the salvage cash inflow is untaxed (no gain/loss)."
ws["D62"].font = small_grey

ws["B64"] = "Units: $000s throughout unless noted. Volume in 000 cases. Prices in $ per case."
ws["B64"].font = small_grey

# ========== 02_Cash_Flows ==========
ws = wb.create_sheet(CF)
ws.sheet_properties.tabColor = "2E75B6"
set_col_widths(ws, [4, 40] + [13] * 9 + [16])
landscape(ws)
ws.freeze_panes = "C6"

ws["B2"] = "Unlevered project cash flows"
ws["B2"].font = title_font
ws["B3"] = "All rows are formulas from 01_Assumptions. FCF = NOPAT + D&A − capex − maintenance − ΔNWC + salvage."
ws["B3"].font = italic_grey

style_header_cell(ws.cell(5, 2), "Line")
for i, y in enumerate(YEARS):
    style_header_cell(ws.cell(5, 3 + i), str(y))

ws["B6"] = "Time index t"
for i in range(N):
    cell = ws.cell(6, 3 + i, f"='{ASSUMP}'!{ycol(i)}34")
    style_formula(cell)
    cell.number_format = "0"

ws["B8"] = "Operating build"
shade_section(ws, 8, until=11)

ws["B9"] = "Volume (000 cases)"
ws["B10"] = "Incremental revenue"
ws["B11"] = "Incremental gross profit"
ws["B12"] = "Cost savings"
ws["B13"] = "Variable distribution"
ws["B14"] = "Incremental G&A"
ws["B15"] = "Incremental EBITDA"
ws["B16"] = "D&A"
ws["B17"] = "EBIT"
ws["B18"] = "Tax on EBIT"
ws["B19"] = "NOPAT"
ws["B20"] = "+ D&A"
ws["B21"] = "Gross cash flow"

for i in range(N):
    c = ycol(i)
    a = f"'{ASSUMP}'!{c}"
    # volume
    cell = ws.cell(9, 3 + i, f"={a}43")
    style_formula(cell)
    cell.number_format = int_fmt
    # revenue = vol * ASP
    cell = ws.cell(10, 3 + i, f"={c}9*'{ASSUMP}'!$C$48")
    style_formula(cell)
    cell.number_format = money
    # GP = vol * contrib * contrib factor
    cell = ws.cell(11, 3 + i, f"={c}9*'{ASSUMP}'!$C$49*'{ASSUMP}'!$C$10")
    style_formula(cell)
    cell.number_format = money
    # savings = run-rate * ramp * factor
    cell = ws.cell(12, 3 + i, f"='{ASSUMP}'!$C$56*{a}36*'{ASSUMP}'!$C$9")
    style_formula(cell)
    cell.number_format = money
    # dist
    cell = ws.cell(13, 3 + i, f"={c}9*'{ASSUMP}'!$C$50")
    style_formula(cell)
    cell.number_format = money
    # G&A
    cell = ws.cell(14, 3 + i, f"={a}44")
    style_formula(cell)
    cell.number_format = money
    # EBITDA
    cell = ws.cell(15, 3 + i, f"={c}11+{c}12-{c}13-{c}14")
    style_formula(cell, key=True)
    cell.number_format = money
    # D&A if 1 <= t <= life
    cell = ws.cell(16, 3 + i, f'=IF(AND({c}6>=1,{c}6<=\'{ASSUMP}\'!$C$27),\'{ASSUMP}\'!$C$62,0)')
    style_formula(cell)
    cell.number_format = money
    # EBIT
    cell = ws.cell(17, 3 + i, f"={c}15-{c}16")
    style_formula(cell)
    cell.number_format = money
    # Tax (allow benefit)
    cell = ws.cell(18, 3 + i, f"={c}17*'{ASSUMP}'!$C$26")
    style_formula(cell)
    cell.number_format = money
    # NOPAT
    cell = ws.cell(19, 3 + i, f"={c}17-{c}18")
    style_formula(cell)
    cell.number_format = money
    # + D&A
    cell = ws.cell(20, 3 + i, f"={c}16")
    style_formula(cell)
    cell.number_format = money
    # GCF
    cell = ws.cell(21, 3 + i, f"={c}19+{c}20")
    style_formula(cell, key=True)
    cell.number_format = money

ws["B23"] = "Investment"
shade_section(ws, 23, until=11)

ws["B24"] = "Gross capex (spend)"
ws["B25"] = "Maintenance capex"
ws["B26"] = "NWC balance"
ws["B27"] = "Δ NWC (use of cash)"
ws["B28"] = "After-tax salvage"
ws["B29"] = "Unlevered FCF"
ws["B30"] = "Cumulative FCF"

for i in range(N):
    c = ycol(i)
    a = f"'{ASSUMP}'!{c}"
    prev = ycol(i - 1) if i else None
    cell = ws.cell(24, 3 + i, f"={a}41")
    style_formula(cell)
    cell.number_format = money
    cell = ws.cell(25, 3 + i, f"={a}42")
    style_formula(cell)
    cell.number_format = money
    # NWC = 0 in final year (t = life) so balance is released
    cell = ws.cell(
        26,
        3 + i,
        f'=IF({c}6>=\'{ASSUMP}\'!$C$27,0,\'{ASSUMP}\'!$C$29*{c}10)',
    )
    style_formula(cell)
    cell.number_format = money
    if i == 0:
        cell = ws.cell(27, 3 + i, f"={c}26")
    else:
        cell = ws.cell(27, 3 + i, f"={c}26-{prev}26")
    style_formula(cell)
    cell.number_format = money
    # salvage in final year t = life
    cell = ws.cell(28, 3 + i, f'=IF({c}6=\'{ASSUMP}\'!$C$27,\'{ASSUMP}\'!$C$60,0)')
    style_formula(cell)
    cell.number_format = money
    # FCF
    cell = ws.cell(29, 3 + i, f"={c}21-{c}24-{c}25-{c}27+{c}28")
    style_formula(cell, key=True)
    cell.number_format = money
    if i == 0:
        cell = ws.cell(30, 3 + i, f"={c}29")
    else:
        cell = ws.cell(30, 3 + i, f"={prev}30+{c}29")
    style_formula(cell, key=True)
    cell.number_format = money

ws["B32"] = "Discounting"
shade_section(ws, 32, until=11)
ws["B33"] = "Discount factor"
ws["B34"] = "PV of FCF"
ws["B35"] = "Cumulative PV of FCF"

for i in range(N):
    c = ycol(i)
    prev = ycol(i - 1) if i else None
    cell = ws.cell(33, 3 + i, f"=1/(1+'{ASSUMP}'!$C$25)^{c}6")
    style_formula(cell)
    cell.number_format = "0.000"
    cell = ws.cell(34, 3 + i, f"={c}29*{c}33")
    style_formula(cell)
    cell.number_format = money
    if i == 0:
        cell = ws.cell(35, 3 + i, f"={c}34")
    else:
        cell = ws.cell(35, 3 + i, f"={prev}35+{c}34")
    style_formula(cell, key=True)
    cell.number_format = money

for r in list(range(6, 7)) + list(range(9, 22)) + list(range(24, 31)) + list(range(33, 36)):
    ws.cell(r, 2).border = thin
    if r in (15, 21, 29, 30, 35):
        ws.cell(r, 2).font = bold

ws["B37"] = (
    "NWC is released in the final project year (balance set to 0). "
    "Salvage equals the residual book value, so it is untaxed. "
    "Tax on EBIT can be negative in the install year (tax shield)."
)
ws["B37"].font = small_grey
ws["B37"].alignment = Alignment(wrap_text=True)
ws.merge_cells("B37:K38")

# ========== 03_Returns ==========
ws = wb.create_sheet(RET)
ws.sheet_properties.tabColor = "548235"
set_col_widths(ws, [4, 44, 16, 16, 16, 16, 16, 16, 16, 16, 16, 22])
landscape(ws)

ws["B2"] = "Returns"
ws["B2"].font = title_font
ws["B3"] = "NPV at WACC, IRR, payback (undiscounted and discounted), profitability index. All formulas from 02_Cash_Flows."
ws["B3"].font = italic_grey

ws["B5"] = "Investment committee metrics"
shade_section(ws, 5, until=5)

metrics = [
    (6, "NPV ($000s)", f"=NPV('{ASSUMP}'!C25,'{CF}'!D29:K29)+'{CF}'!C29", money, True),
    (7, "IRR", f"=IRR('{CF}'!C29:K29)", pct, True),
    (8, "WACC (hurdle)", f"='{ASSUMP}'!C25", pct, False),
    (9, "IRR − WACC (spread)", "=C7-C8", pct, True),
    (10, "Total gross capex ($000s)", f"='{ASSUMP}'!C59", money, False),
    (11, "PV of capex + maintenance ($000s)", f"=SUMPRODUCT('{CF}'!C24:K24+'{CF}'!C25:K25,'{CF}'!C33:K33)", money, False),
    (12, "Profitability index", "=IF(C11=0,\"n.m.\",(C6+C11)/C11)", "0.00", True),
    (13, "Undiscounted payback (years from 2026)", "=C42", years_fmt, True),
    (14, "Discounted payback (years from 2026)", "=C48", years_fmt, True),
    (15, "Sum of unlevered FCF ($000s)", f"=SUM('{CF}'!C29:K29)", money, False),
]
for r, lab, formula, fmt, key in metrics:
    ws.cell(r, 2, lab).border = thin
    ws.cell(r, 2).font = bold if key else black
    cell = ws.cell(r, 3, formula)
    style_formula(cell, key=key)
    cell.number_format = fmt

ws["D6"] = "Year-end discounting. NPV = FCF₀ + NPV(WACC, FCF₁…FCF₈)."
ws["D6"].font = small_grey
ws["D7"] = "IRR on the unlevered FCF series including t = 0 spend."
ws["D7"].font = small_grey
ws["D12"] = "(NPV + PV of investment) / PV of investment. >1 with NPV > 0."
ws["D12"].font = small_grey

ws["B17"] = "Recommendation"
shade_section(ws, 17, until=5)
ws["B18"] = "Decision"
ws["C18"] = '=IF(AND(ISNUMBER(C6),C6>0,ISNUMBER(C7),C7>C8,ISNUMBER(C13),C13<=5),"Recommend approve","Review — NPV, spread, or payback off-hurdle")'
style_formula(ws["C18"], key=True)
ws["C18"].alignment = Alignment(horizontal="left")
ws.merge_cells("C18:G18")
ws["B19"] = "Rule"
ws["C19"] = "Approve if NPV > 0, IRR > WACC, and undiscounted payback ≤ 5 years. Otherwise review."
ws["C19"].font = small_grey
ws.merge_cells("C19:G19")

ws["C6"].fill = green_fill
ws["C7"].fill = green_fill

ws.conditional_formatting.add(
    "C6",
    CellIsRule(operator="lessThan", formula=["0"], fill=red_fill),
)
ws.conditional_formatting.add(
    "C9",
    CellIsRule(operator="lessThan", formula=["0"], fill=red_fill),
)

# FCF strip for charts + payback math
ws["B21"] = "Cash-flow strip (for payback and chart)"
shade_section(ws, 21, until=11)
style_header_cell(ws.cell(22, 2), "Item")
for i, y in enumerate(YEARS):
    style_header_cell(ws.cell(22, 3 + i), str(y))

ws["B23"] = "t"
ws["B24"] = "Unlevered FCF"
ws["B25"] = "Cumulative FCF"
ws["B26"] = "PV of FCF"
ws["B27"] = "Cumulative PV"
ws["B28"] = "Payback crossing (undiscounted)"
ws["B29"] = "Payback crossing (discounted)"

for i in range(N):
    c = ycol(i)
    prev = ycol(i - 1) if i else None
    cell = ws.cell(23, 3 + i, f"='{CF}'!{c}6")
    style_formula(cell)
    cell.number_format = "0"
    cell = ws.cell(24, 3 + i, f"='{CF}'!{c}29")
    style_formula(cell, key=True)
    cell.number_format = money
    cell = ws.cell(25, 3 + i, f"='{CF}'!{c}30")
    style_formula(cell)
    cell.number_format = money
    cell = ws.cell(26, 3 + i, f"='{CF}'!{c}34")
    style_formula(cell)
    cell.number_format = money
    cell = ws.cell(27, 3 + i, f"='{CF}'!{c}35")
    style_formula(cell)
    cell.number_format = money
    # crossing contribution: if prior cum < 0 and this cum >= 0, prior_t + abs(prior_cum)/this_fcf
    if i == 0:
        cell = ws.cell(28, 3 + i, f'=IF({c}25>=0,0,0)')
        cell2 = ws.cell(29, 3 + i, f'=IF({c}27>=0,0,0)')
    else:
        cell = ws.cell(
            28,
            3 + i,
            f'=IF(AND({prev}25<0,{c}25>=0),{prev}23+ABS({prev}25)/{c}24,0)',
        )
        cell2 = ws.cell(
            29,
            3 + i,
            f'=IF(AND({prev}27<0,{c}27>=0),{prev}23+ABS({prev}27)/{c}26,0)',
        )
    style_formula(cell)
    cell.number_format = years_fmt
    style_formula(cell2)
    cell2.number_format = years_fmt

for r in range(23, 30):
    ws.cell(r, 2).border = thin

ws["B31"] = "Payback helpers"
shade_section(ws, 31, until=5)
ws["B32"] = "First year cumulative FCF ≥ 0 (t)"
ws["C32"] = '=IF(COUNTIF(C25:K25,">=0")=0,"n.m.",INDEX(C23:K23,1,MATCH(TRUE,INDEX(C25:K25>=0,0),0)))'
style_formula(ws["C32"])
ws["C32"].number_format = "0"
ws["B33"] = "Crossing flag sum (undiscounted)"
ws["C33"] = "=SUM(C28:K28)"
style_formula(ws["C33"])
ws["C33"].number_format = years_fmt
ws["B34"] = "Crossing flag sum (discounted)"
ws["C34"] = "=SUM(C29:K29)"
style_formula(ws["C34"])
ws["C34"].number_format = years_fmt

ws["B36"] = "Undiscounted payback (years)"
ws["C36"] = '=IF(K25<0,NA(),IF(C25>=0,0,C33))'
style_formula(ws["C36"], key=True)
ws["C36"].number_format = years_fmt
ws["B37"] = "Discounted payback (years)"
ws["C37"] = '=IF(K27<0,NA(),IF(C27>=0,0,C34))'
style_formula(ws["C37"], key=True)
ws["C37"].number_format = years_fmt

# numeric aliases used by the summary (C13/C14)
ws["C42"] = "=C36"
ws["C48"] = "=C37"
ws["C42"].number_format = years_fmt
ws["C48"].number_format = years_fmt
ws["C42"].font = black
ws["C48"].font = black
# hide aliases off to the right of the used print area conceptually — leave visible but unlabeled is messy.
# Put them next to the helpers instead; C13/C14 already formula-link to C42/C48.
ws["B42"] = "(alias for C13)"
ws["B42"].font = small_grey
ws["B48"] = "(alias for C14)"
ws["B48"].font = small_grey

ws["B39"] = "Annual unlevered FCF and cumulative ($000s)"
ws["B39"].font = section_font

chart = BarChart()
chart.type = "col"
chart.grouping = "clustered"
chart.title = "Annual unlevered FCF ($000s)"
chart.y_axis.title = "$000s"
chart.height = 8
chart.width = 16
chart.legend.position = "b"
data = Reference(ws, min_col=3, min_row=24, max_col=11, max_row=24)
cats = Reference(ws, min_col=3, min_row=22, max_col=11)
chart.add_data(data, from_rows=True, titles_from_data=False)
chart.set_categories(cats)
chart.shape = 4
ws.add_chart(chart, "B50")

chart2 = LineChart()
chart2.title = "Cumulative unlevered FCF ($000s)"
chart2.y_axis.title = "$000s"
chart2.height = 8
chart2.width = 16
chart2.legend.position = "b"
data2 = Reference(ws, min_col=3, min_row=25, max_col=11, max_row=25)
chart2.add_data(data2, from_rows=True, titles_from_data=False)
chart2.set_categories(cats)
ws.add_chart(chart2, "B66")

ws["B82"] = "Payback is interpolated in the crossing year: t_prior + |cumulative_prior| / FCF_crossing. n.m. if cumulative never turns positive."
ws["B82"].font = small_grey
ws["B82"].alignment = Alignment(wrap_text=True)

# ========== 04_Sensitivity ==========
ws = wb.create_sheet(SENS)
ws.sheet_properties.tabColor = "C65911"
set_col_widths(ws, [4, 28] + [14] * 10)
landscape(ws)

ws["B2"] = "Sensitivity"
ws["B2"].font = title_font
ws["B3"] = "WACC one-way uses the live FCF strip. Volume two-way rebuilds FCF (volume-driven lines × factor; savings, capex, G&A, D&A held at the current case)."
ws["B3"].font = italic_grey

ws["B5"] = "Base case (current scenario)"
shade_section(ws, 5, until=5)
ws["B6"] = "NPV"
ws["C6"] = f"='{RET}'!C6"
style_formula(ws["C6"], key=True)
ws["C6"].number_format = money
ws["B7"] = "IRR"
ws["C7"] = f"='{RET}'!C7"
style_formula(ws["C7"], key=True)
ws["C7"].number_format = pct
ws["B8"] = "WACC"
ws["C8"] = f"='{ASSUMP}'!C25"
style_formula(ws["C8"])
ws["C8"].number_format = pct

# One-way WACC
ws["B10"] = "One-way — NPV vs WACC (live FCF strip)"
shade_section(ws, 10, until=8)
style_header_cell(ws.cell(11, 2), "WACC")
style_header_cell(ws.cell(11, 3), "NPV ($000s)")
style_header_cell(ws.cell(11, 4), "vs base")
waccs = [0.070, 0.080, 0.095, 0.110, 0.125]
for i, w in enumerate(waccs):
    r = 12 + i
    cell = ws.cell(r, 2, w)
    style_input(cell)
    cell.number_format = pct
    cell = ws.cell(r, 3, f"=NPV(B{r},'{CF}'!D29:K29)+'{CF}'!C29")
    style_formula(cell, key=(abs(w - 0.095) < 1e-9))
    cell.number_format = money
    cell = ws.cell(r, 4, f"=C{r}-$C$6")
    style_formula(cell)
    cell.number_format = money

# Two-way engine: FCF by volume factor
# Building blocks from CF, scale GP, dist, revenue/NWC, leave savings/G&A/D&A/capex
ws["B19"] = "Two-way engine — FCF by volume factor (current case = 100%)"
shade_section(ws, 19, until=11)
style_header_cell(ws.cell(20, 2), "Volume factor \\ Year")
for i, y in enumerate(YEARS):
    style_header_cell(ws.cell(20, 3 + i), str(y))

vol_factors = [0.80, 0.90, 1.00, 1.10, 1.20]
# Rows 21-25: FCF for each volume factor
for vi, vf in enumerate(vol_factors):
    r = 21 + vi
    cell = ws.cell(r, 2, vf)
    style_input(cell)
    cell.number_format = pct
    for i in range(N):
        c = ycol(i)
        # FCF(v) = [v*GP + Sav - v*Dist - G&A - D&A]*(1-t) + D&A - Capex - Maint - ΔNWC(v) + Salv
        # NWC_t(v) = IF(t>=life, 0, nwc% * v * rev)
        # ΔNWC from NWC path
        # Implement NWC via nested IFs using CF revenue and t
        # NWC_t = IF(t>=life,0, nwc% * vf * revenue_t)
        prev = ycol(i - 1)
        if i == 0:
            dnwc = f"IF('{CF}'!{c}6>='{ASSUMP}'!$C$27,0,'{ASSUMP}'!$C$29*$B{r}*'{CF}'!{c}10)"
        else:
            nwc_t = f"IF('{CF}'!{c}6>='{ASSUMP}'!$C$27,0,'{ASSUMP}'!$C$29*$B{r}*'{CF}'!{c}10)"
            nwc_p = f"IF('{CF}'!{prev}6>='{ASSUMP}'!$C$27,0,'{ASSUMP}'!$C$29*$B{r}*'{CF}'!{prev}10)"
            dnwc = f"({nwc_t})-({nwc_p})"
        ebit = (
            f"($B{r}*'{CF}'!{c}11+'{CF}'!{c}12-$B{r}*'{CF}'!{c}13-'{CF}'!{c}14-'{CF}'!{c}16)"
        )
        fcf = (
            f"={ebit}*(1-'{ASSUMP}'!$C$26)+'{CF}'!{c}16-'{CF}'!{c}24-'{CF}'!{c}25-({dnwc})+'{CF}'!{c}28"
        )
        cell = ws.cell(r, 3 + i, fcf)
        style_formula(cell, key=(abs(vf - 1.0) < 1e-9))
        cell.number_format = money

# NPV grid WACC (rows) × volume (cols)
ws["B28"] = "Two-way — NPV ($000s): rows = WACC, columns = volume factor"
shade_section(ws, 28, until=8)
ws["B29"] = "WACC \\ volume"
ws["B29"].font = bold
ws["B29"].fill = section_fill
ws["B29"].border = thin
for i, vf in enumerate(vol_factors):
    cell = ws.cell(29, 3 + i, f"=B{21+i}")
    style_formula(cell)
    cell.number_format = pct
    cell.fill = section_fill
    cell.font = bold

grid_waccs = [0.070, 0.080, 0.095, 0.110, 0.125]
for ri, w in enumerate(grid_waccs):
    r = 30 + ri
    cell = ws.cell(r, 2, w)
    style_input(cell)
    cell.number_format = pct
    cell.fill = section_fill
    for ci in range(5):
        frow = 21 + ci
        # NPV = C_frow + NPV(wacc, D:K of that row)
        cell = ws.cell(r, 3 + ci, f"=NPV($B{r},{get_column_letter(4)}{frow}:K{frow})+C{frow}")
        style_formula(cell, key=(abs(w - 0.095) < 1e-9 and ci == 2))
        cell.number_format = money

# Highlight the 100% / 9.5% cell already keyed.
# Color scale-ish via above/below base
ws.conditional_formatting.add(
    "C30:G34",
    FormulaRule(formula=["C30>=$C$6"], fill=green_fill),
)
ws.conditional_formatting.add(
    "C30:G34",
    FormulaRule(formula=["C30<$C$6"], fill=amber_fill),
)

# Tornado
ws["B37"] = "Tornado — NPV vs driver (low / high vs current case)"
shade_section(ws, 37, until=8)
headers = ["Driver", "Low", "High", "NPV low", "NPV high", "Swing (high − low)"]
for i, h in enumerate(headers):
    style_header_cell(ws.cell(38, 2 + i), h)

# Row 39 WACC 8% / 11%
ws["B39"] = "WACC"
ws["C39"] = 0.08
ws["D39"] = 0.11
style_input(ws["C39"])
style_input(ws["D39"])
ws["C39"].number_format = pct
ws["D39"].number_format = pct
ws["E39"] = f"=NPV(C39,'{CF}'!D29:K29)+'{CF}'!C29"
ws["F39"] = f"=NPV(D39,'{CF}'!D29:K29)+'{CF}'!C29"
style_formula(ws["E39"])
style_formula(ws["F39"])
ws["E39"].number_format = money
ws["F39"].number_format = money
ws["G39"] = "=F39-E39"
style_formula(ws["G39"], key=True)
ws["G39"].number_format = money

# Row 40 volume 80% / 120% — use engine rows 21 and 25 at live WACC
ws["B40"] = "Volume factor"
ws["C40"] = "=B21"
ws["D40"] = "=B25"
style_formula(ws["C40"])
style_formula(ws["D40"])
ws["C40"].number_format = pct
ws["D40"].number_format = pct
ws["E40"] = f"=NPV('{ASSUMP}'!C25,D21:K21)+C21"
ws["F40"] = f"=NPV('{ASSUMP}'!C25,D25:K25)+C25"
style_formula(ws["E40"])
style_formula(ws["F40"])
ws["E40"].number_format = money
ws["F40"].number_format = money
ws["G40"] = "=F40-E40"
style_formula(ws["G40"], key=True)
ws["G40"].number_format = money

# Row 41 capex factor 0.90 / 1.15 around current
# Approximate: extra capex = (k-1)*gross capex path, extra D&A = (k-1)*D&A, extra salvage = (k-1)*salvage
# FCF_k ≈ FCF + (1-k)*capex - (k-1)*ΔD&A_tax_effect + (k-1)*salvage
# D&A scales with capex; NOPAT includes -D&A*(1-t) + D&A = D&A*t tax shield
# FCF extra from capex scale k vs 1:
#   -(k-1)*capex + (k-1)*salvage + (k-1)*D&A*tax
# NPV extra = that series discounted
ws["B41"] = "Capex factor (vs current)"
ws["C41"] = 0.90
ws["D41"] = 1.15
style_input(ws["C41"])
style_input(ws["D41"])
ws["C41"].number_format = "0.00"
ws["D41"].number_format = "0.00"
# NPV at k: base NPV + (k-1) * PV(-(capex) + salvage + D&A*tax)
# PV of capex already in RET C11 which includes maint — use capex only + salvage + D&A tax shield
ws["E41"] = (
    f"='{RET}'!C6+(C41-1)*("
    f"SUMPRODUCT(-'{CF}'!C24:K24+'{CF}'!C28:K28+'{CF}'!C16:K16*'{ASSUMP}'!$C$26,'{CF}'!C33:K33)"
    f")"
)
ws["F41"] = (
    f"='{RET}'!C6+(D41-1)*("
    f"SUMPRODUCT(-'{CF}'!C24:K24+'{CF}'!C28:K28+'{CF}'!C16:K16*'{ASSUMP}'!$C$26,'{CF}'!C33:K33)"
    f")"
)
style_formula(ws["E41"])
style_formula(ws["F41"])
ws["E41"].number_format = money
ws["F41"].number_format = money
ws["G41"] = "=F41-E41"
style_formula(ws["G41"], key=True)
ws["G41"].number_format = money

# Row 42 savings 80% / 120%
# Savings is CF row 12; after tax because it sits in EBIT
ws["B42"] = "Savings factor (vs current)"
ws["C42"] = 0.80
ws["D42"] = 1.20
style_input(ws["C42"])
style_input(ws["D42"])
ws["C42"].number_format = pct
ws["D42"].number_format = pct
ws["E42"] = (
    f"='{RET}'!C6+(C42-1)*SUMPRODUCT('{CF}'!C12:K12*(1-'{ASSUMP}'!$C$26),'{CF}'!C33:K33)"
)
ws["F42"] = (
    f"='{RET}'!C6+(D42-1)*SUMPRODUCT('{CF}'!C12:K12*(1-'{ASSUMP}'!$C$26),'{CF}'!C33:K33)"
)
style_formula(ws["E42"])
style_formula(ws["F42"])
ws["E42"].number_format = money
ws["F42"].number_format = money
ws["G42"] = "=F42-E42"
style_formula(ws["G42"], key=True)
ws["G42"].number_format = money

# Row 43 contribution $/case  -15% / +10%
ws["B43"] = "Contribution $/case (vs current)"
ws["C43"] = 0.85
ws["D43"] = 1.10
style_input(ws["C43"])
style_input(ws["D43"])
ws["C43"].number_format = pct
ws["D43"].number_format = pct
ws["E43"] = (
    f"='{RET}'!C6+(C43-1)*SUMPRODUCT('{CF}'!C11:K11*(1-'{ASSUMP}'!$C$26),'{CF}'!C33:K33)"
)
ws["F43"] = (
    f"='{RET}'!C6+(D43-1)*SUMPRODUCT('{CF}'!C11:K11*(1-'{ASSUMP}'!$C$26),'{CF}'!C33:K33)"
)
style_formula(ws["E43"])
style_formula(ws["F43"])
ws["E43"].number_format = money
ws["F43"].number_format = money
ws["G43"] = "=F43-E43"
style_formula(ws["G43"], key=True)
ws["G43"].number_format = money

for r in range(39, 44):
    ws.cell(r, 2).border = thin

ws["B45"] = (
    "Volume two-way rebuilds GP, distribution, and NWC with the factor; savings / G&A / D&A / capex stay at the current scenario. "
    "Tornado capex, savings, and contribution rows are first-order (scale that line, hold the rest)."
)
ws["B45"].font = small_grey
ws["B45"].alignment = Alignment(wrap_text=True)
ws.merge_cells("B45:G46")

# ========== 05_One_Pager ==========
ws = wb.create_sheet(ONE)
ws.sheet_properties.tabColor = "1F4E79"
set_col_widths(ws, [4, 22, 18, 4, 22, 18, 4, 24, 18, 4, 20])
landscape(ws, fit_height=1)

ws["B2"] = "Northline Consumer Products"
ws["B2"].font = title_font
ws["B3"] = f"='{ASSUMP}'!C20"
ws["B3"].font = section_font
ws.merge_cells("B3:I3")
ws["B4"] = "Investment committee one-pager  ·  unlevered  ·  $000s  ·  fictional sample"
ws["B4"].font = italic_grey

ws["B6"] = "Snapshot"
shade_section(ws, 6, until=9)

def tile(ws, r, c, label, formula, fmt):
    lab = ws.cell(r, c, label)
    lab.font = tile_label
    lab.fill = tile_fill
    lab.border = thin
    val = ws.cell(r + 1, c, formula)
    val.font = tile_value
    val.fill = tile_fill
    val.border = thin
    val.number_format = fmt
    val.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 1)
    ws.merge_cells(start_row=r + 1, start_column=c, end_row=r + 1, end_column=c + 1)
    lab.alignment = Alignment(horizontal="center")

tile(ws, 7, 2, "NPV ($000s)", f"='{RET}'!C6", money)
tile(ws, 7, 5, "IRR", f"='{RET}'!C7", pct)
tile(ws, 7, 8, "Payback (years)", f"='{RET}'!C13", years_fmt)
tile(ws, 10, 2, "WACC", f"='{ASSUMP}'!C25", pct)
tile(ws, 10, 5, "Profitability index", f"='{RET}'!C12", "0.00")
tile(ws, 10, 8, "Total capex ($000s)", f"='{ASSUMP}'!C59", money)

ws["B13"] = "Recommendation"
shade_section(ws, 13, until=9)
ws["B14"] = f"='{RET}'!C18"
ws["B14"].font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
ws["B14"].fill = green_fill
ws["B14"].border = thin
ws.merge_cells("B14:I14")
ws["B15"] = f"='{RET}'!C19"
ws["B15"].font = small_grey
ws.merge_cells("B15:I15")

ws["B17"] = "Key assumptions (live)"
shade_section(ws, 17, until=9)
labels = [
    (18, "Scenario", f"='{ASSUMP}'!D6", "@"),
    (19, "WACC / tax / life", f"=TEXT('{ASSUMP}'!C25,\"0.0%\")&\"  /  \"&TEXT('{ASSUMP}'!C26,\"0.0%\")&\"  /  \"&'{ASSUMP}'!C27&\" yrs\"", "@"),
    (20, "Gross capex 2026–27 ($000s)", f"='{ASSUMP}'!C41&\" in 2026,  \"&'{ASSUMP}'!D41&\" in 2027\"", "@"),
    (21, "Run-rate volume (000 cases)", f"='{ASSUMP}'!C47*'{ASSUMP}'!C8", int_fmt),
    (22, "Contribution $/case (scenario)", f"='{ASSUMP}'!C49*'{ASSUMP}'!C10", ppu),
    (23, "Run-rate savings ($000s)", f"='{ASSUMP}'!C56*'{ASSUMP}'!C9", money),
]
ws["C20"] = f"='{ASSUMP}'!C41"
# redo row 20 as two numeric pulls instead of concatenation of numbers without format
for r, lab, formula, fmt in [
    (18, "Scenario", f"='{ASSUMP}'!D6", "@"),
    (19, "WACC", f"='{ASSUMP}'!C25", pct),
    (20, "Tax rate", f"='{ASSUMP}'!C26", pct),
    (21, "Project life (years)", f"='{ASSUMP}'!C27", "0"),
    (22, "Gross capex ($000s)", f"='{ASSUMP}'!C59", money),
    (23, "Run-rate volume (000 cases)", f"='{ASSUMP}'!C43", int_fmt),
    (24, "Contribution $/case", f"='{ASSUMP}'!C49*'{ASSUMP}'!C10", ppu),
    (25, "Run-rate savings ($000s)", f"='{ASSUMP}'!C56*'{ASSUMP}'!C9", money),
    (26, "NWC % of revenue", f"='{ASSUMP}'!C29", pct),
    (27, "Salvage ($000s)", f"='{ASSUMP}'!C60", money),
]:
    ws.cell(r, 2, lab).border = thin
    cell = ws.cell(r, 3, formula)
    style_formula(cell)
    cell.number_format = fmt

# C23 volume at run-rate after factor: C47*C8 is cleaner than C43 which is 2034
ws["C23"] = f"='{ASSUMP}'!C47*'{ASSUMP}'!C8"
style_formula(ws["C23"])
ws["C23"].number_format = int_fmt

ws["E18"] = "IRR − WACC"
ws["E18"].border = thin
ws["F18"] = f"='{RET}'!C9"
style_formula(ws["F18"], key=True)
ws["F18"].number_format = pct
ws["E19"] = "Discounted payback (yrs)"
ws["E19"].border = thin
ws["F19"] = f"='{RET}'!C14"
style_formula(ws["F19"])
ws["F19"].number_format = years_fmt
ws["E20"] = "Sum of FCF ($000s)"
ws["E20"].border = thin
ws["F20"] = f"='{RET}'!C15"
style_formula(ws["F20"])
ws["F20"].number_format = money
ws["E21"] = "PV of investment ($000s)"
ws["E21"].border = thin
ws["F21"] = f"='{RET}'!C11"
style_formula(ws["F21"])
ws["F21"].number_format = money

ws["B29"] = "Unlevered FCF ($000s)"
ws["B29"].font = section_font

# Mini FCF row for the one-pager chart
style_header_cell(ws.cell(30, 2), "Year")
for i, y in enumerate(YEARS):
    style_header_cell(ws.cell(30, 3 + i), str(y))
ws["B31"] = "FCF"
for i in range(N):
    cell = ws.cell(31, 3 + i, f"='{CF}'!{ycol(i)}29")
    style_formula(cell, key=True)
    cell.number_format = money
ws["B32"] = "Cumulative"
for i in range(N):
    cell = ws.cell(32, 3 + i, f"='{CF}'!{ycol(i)}30")
    style_formula(cell)
    cell.number_format = money

chart = BarChart()
chart.type = "col"
chart.title = "Annual unlevered FCF ($000s)"
chart.y_axis.title = "$000s"
chart.height = 7
chart.width = 15
chart.legend = None
data = Reference(ws, min_col=3, min_row=31, max_col=11, max_row=31)
cats = Reference(ws, min_col=3, min_row=30, max_col=11)
chart.add_data(data, from_rows=True, titles_from_data=False)
chart.set_categories(cats)
ws.add_chart(chart, "B34")

ws["B49"] = "NPV vs WACC (from 04_Sensitivity)"
ws["B49"].font = section_font
ws["B50"] = f"='{SENS}'!B12"
ws["C50"] = f"='{SENS}'!C12"
ws["B51"] = f"='{SENS}'!B13"
ws["C51"] = f"='{SENS}'!C13"
ws["B52"] = f"='{SENS}'!B14"
ws["C52"] = f"='{SENS}'!C14"
ws["B53"] = f"='{SENS}'!B15"
ws["C53"] = f"='{SENS}'!C15"
ws["B54"] = f"='{SENS}'!B16"
ws["C54"] = f"='{SENS}'!C16"
for r in range(50, 55):
    style_formula(ws.cell(r, 2))
    ws.cell(r, 2).number_format = pct
    style_formula(ws.cell(r, 3))
    ws.cell(r, 3).number_format = money
    ws.cell(r, 2).border = thin

ws["E49"] = "How to inherit"
ws["E49"].font = section_font
ws["E50"] = "Change yellow cells on 01_Assumptions. Live case is whatever scenario 1/2/3 is selected. Do not type over black cells."
ws["E50"].alignment = Alignment(wrap_text=True)
ws.merge_cells("E50:I52")
ws["E50"].font = italic_grey

ws["B56"] = "All sample numbers are fictional. Built for a public GitHub portfolio — no employer data."
ws["B56"].font = small_grey

# ========== 06_Data_Dictionary ==========
ws = wb.create_sheet("06_Data_Dictionary")
ws.sheet_properties.tabColor = "7F7F7F"
set_col_widths(ws, [4, 36, 22, 82])
landscape(ws)
ws["B2"] = "Data dictionary"
ws["B2"].font = title_font
ws["B3"] = "Field definitions so another analyst can inherit the file."
ws["B3"].font = italic_grey
headers = ["Field", "Tab", "Definition"]
for i, h in enumerate(headers):
    cell = ws.cell(5, 2 + i, h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin

defs = [
    ("Scenario", ASSUMP, "1 Base / 2 Upside / 3 Downside. Indexes the factor table and scales capex, volume, savings, and contribution $/case."),
    ("Capex / volume / savings / contribution factors", ASSUMP, "Applied to gross capex spend, volume, run-rate savings, and contribution per case."),
    ("WACC", ASSUMP, "Unlevered project hurdle. Year-end discounting of FCF. Sample 9.5%."),
    ("Tax rate", ASSUMP, "Applied to EBIT. Negative EBIT produces a tax shield (assumes the firm has other taxable income)."),
    ("Project life", ASSUMP, "Operating and D&A years from COD (t = 1). Sample 8 years (2027–2034). t = 0 is 2026 spend."),
    ("Salvage % of gross capex", ASSUMP, "Cash inflow in the final year. Straight-line D&A is charged on (gross capex − salvage), so book value = salvage and the inflow is untaxed."),
    ("NWC % of incremental revenue", ASSUMP, "NWC balance = % × incremental revenue. Released (balance → 0) in the final project year."),
    ("Gross capex timing", ASSUMP, "Equipment, installation, and spares by year, then × capex factor. Sample spend in 2026–2027."),
    ("Maintenance capex", ASSUMP, "Sustaining capex after COD. Not scaled by the scenario capex factor."),
    ("Volume ramp", ASSUMP, "Share of run-rate volume lift by year, then × volume factor."),
    ("Savings ramp", ASSUMP, "Share of run-rate cost savings by year, then × savings factor."),
    ("Run-rate volume lift", ASSUMP, "Incremental 000 cases vs doing nothing once the line is at nameplate."),
    ("Incremental contribution $/case", ASSUMP, "Net price − variable COGS on incremental volume. Then × contribution factor."),
    ("Run-rate savings", ASSUMP, "Labor, scrap/giveaway, energy, avoided old-line maintenance at nameplate ($000s)."),
    ("Incremental EBITDA", CF, "Contribution + savings − variable distribution − incremental G&A."),
    ("D&A", CF, "Straight-line annual charge for t = 1..life. Zero in the 2026 spend year."),
    ("NOPAT", CF, "EBIT × (1 − tax)."),
    ("Unlevered FCF", CF, "NOPAT + D&A − gross capex − maintenance − ΔNWC + after-tax salvage."),
    ("NPV", RET, "FCF₀ + Excel NPV(WACC, FCF₁…FCF₈). Year-end convention."),
    ("IRR", RET, "Internal rate of return on the unlevered FCF series including t = 0."),
    ("Payback", RET, "Interpolated year when cumulative FCF turns positive: t_prior + |cum_prior| / FCF_crossing."),
    ("Discounted payback", RET, "Same interpolation on cumulative PV of FCF."),
    ("Profitability index", RET, "(NPV + PV of capex and maintenance) / PV of capex and maintenance."),
    ("Recommendation", RET, "Recommend approve if NPV > 0, IRR > WACC, and undiscounted payback ≤ 5 years."),
    ("Volume two-way", SENS, "Rebuilds FCF with GP, distribution, and NWC × volume factor; other lines held at the current case."),
    ("Tornado", SENS, "One-at-a-time NPV at low/high WACC, volume, capex, savings, and contribution."),
    ("Units", "All", "Dollars in $000s. Volume in 000 cases. Prices and contribution in $ per case. Rates in %."),
]
for i, (field, tab, definition) in enumerate(defs):
    r = 6 + i
    ws.cell(r, 2, field).font = bold
    ws.cell(r, 2).border = thin
    ws.cell(r, 3, tab).border = thin
    cell = ws.cell(r, 4, definition)
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    cell.border = thin
    ws.row_dimensions[r].height = 32

ws.cell(6 + len(defs), 2, "All sample numbers are fictional. Built for a public GitHub portfolio — no employer data.").font = small_grey

# Python-side sanity (does not need Excel)
capex_pf = [e + i + s for e, i, s in zip(equip, install, spares)]  # pre-factor
wacc = 0.095
tax = 0.25
life = 8
salv_pct = 0.08
nwc_pct = 0.065
asp = 11.75
contrib = 3.20
dist = 0.42
run_vol = 1500
run_sav = 2500 + 800 + 400 + 250
ga = 80
vol = [run_vol * r for r in vol_ramp]
sav = [run_sav * r for r in sav_ramp]
total_capex = sum(capex_pf)
salvage = total_capex * salv_pct
da_ann = (total_capex - salvage) / life
fcf = []
nwc_prev = 0.0
cum = 0.0
for t in range(N):
    gp = vol[t] * contrib
    revenue = vol[t] * asp
    ebitda = gp + sav[t] - vol[t] * dist - (ga if t >= 1 else 0)
    da = da_ann if 1 <= t <= life else 0.0
    ebit = ebitda - da
    nopat = ebit * (1 - tax)
    nwc = 0.0 if t >= life else nwc_pct * revenue
    dnwc = nwc - nwc_prev
    nwc_prev = nwc
    salv = salvage if t == life else 0.0
    cf = nopat + da - capex_pf[t] - maint[t] - dnwc + salv
    fcf.append(cf)
    cum += cf
npv = sum(cf / (1 + wacc) ** t for t, cf in enumerate(fcf))
print("Python check total capex $000s", round(total_capex))
print("Python check run-rate EBITDA $000s", round(1500 * 3.20 + run_sav - 1500 * 0.42 - 80))
print("Python check FCF", [round(x) for x in fcf])
print("Python check NPV @ 9.5%", round(npv))
print("Python check sum FCF", round(sum(fcf)))

out = Path(__file__).resolve().parent / "Northline_Capex_Investment_Case.xlsx"
wb.save(out)
print("Wrote", out)
print("Sheets:", wb.sheetnames)
